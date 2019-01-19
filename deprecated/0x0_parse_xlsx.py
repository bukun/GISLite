from openpyxl import load_workbook

wb=load_workbook(filename = 'meta_poly.xlsx')
sheet = wb.active


max_row_num = sheet.max_row
print(max_row_num)

def hex2dec(string_num):
    return str(int(string_num.upper(), 16))

def get_sub_val(idx_row, idx_col):
    print('=====: {}, {}'.format(idx_row, idx_col))
    the_key = sheet.cell(row = idx_row, column = idx_col).value
    test_val = sheet.cell(row = idx_row, column = idx_col + 1)
    if test_val and test_val.value:


        out_val= test_val.value



        print('-' * 40)
        # print(test_val.value)
        return  out_val
    else:
        # 从下一行开始
        print('-' * 40)
        out_dict = {}
        idx_col2 = idx_col + 1
        for xx in range(idx_row + 1, max_row_num + 1):
            val = sheet.cell(row = xx, column = idx_col2 )
            if val and val.value:
                tt = get_sub_val(xx, idx_col2 )
                out_dict[val.value] = tt
        print(out_dict)
        return out_dict
# import collections

the_dic = {}
for x in range(1, max_row_num + 1):
    idx_col = 1
    val = sheet.cell(row = x, column = idx_col)
    if val and val.value:
        tt = get_sub_val(x, idx_col)
        print(val.value)
        the_dic[val.value] = tt
from pprint import pprint

pprint(the_dic)

# ws1=wb.get_sheet_by_name("Sheet1")