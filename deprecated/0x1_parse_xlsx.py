'''
对xlsx进行解析，使用迭代，生成字典结果。
'''
from openpyxl import load_workbook

wb = load_workbook(filename='meta_poly.xlsx')
sheet = wb.active

max_row_num = sheet.max_row
print(max_row_num)


def hex2dec(string_num):
    return str(int(string_num.upper(), 16))

def get_sub_val(idx_row, idx_col, max_row_idx=max_row_num):
    print('x' * 4)
    print(idx_row, idx_col, max_row_idx)
    print('x' * 4)
    the_key = sheet.cell(row=idx_row, column=idx_col).value
    test_val = sheet.cell(row=idx_row, column=idx_col + 1)
    if test_val and test_val.value:
        if 'color' in the_key.lower():
            colors = test_val.fill.fgColor.index
            # print(colors)
            out_val = [int(hex2dec(colors[2:4])), int(hex2dec(colors[4:6])), int(hex2dec(colors[6:8]))]

        else:
            out_val = test_val.value

        # print('-' * 40)
        # print(test_val.value)
        return {the_key: out_val}
    else:
        the_key_arr_indx = []
        for x in range(idx_row, max_row_idx + 1):
            the_ce_val = sheet.cell(row=x, column=idx_col)
            if the_ce_val and the_ce_val.value:
                the_key_arr_indx.append(x)

        the_key_arr_indx.append(max_row_idx)
        print('    ', '=' * 10)
        print('    ', the_key_arr_indx)
        # uu_dict = {}
        out_dict = {}
        for xx_uav, yy_zvl in zip(the_key_arr_indx[:-1], the_key_arr_indx[1:]):
            print('        ', '-' * 4)
            print('        ',xx_uav, yy_zvl)
            # 从下一行开始
            # print('-' * 40)
            yy_key = sheet.cell(row=xx_uav, column=idx_col)
            if yy_key and yy_key.value:
                yy_key = yy_key.value


            idx_col2 = idx_col + 1
            for xxx in range(xx_uav, yy_zvl):
                # print('sub:{}'.format(xxx))
                val = sheet.cell(row=xxx, column=idx_col2)
                if val and val.value:
                    tt = get_sub_val(xxx, idx_col2, yy_zvl)
                    out_dict = dict(out_dict, **tt)
            # uu_dict = dict(uu_dict, **out_dict)

        # print(out_dict)
        # return {the_key: uu_dict}
        return {the_key: out_dict}



the_dic = get_sub_val(1, 1)

from pprint import pprint

pprint(the_dic)

# ws1=wb.get_sheet_by_name("Sheet1")
