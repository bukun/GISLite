# -*- coding: utf-8 -*-

'''
Helper for GISLite.
'''

import os

import yaml
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from jinja2 import Environment, FileSystemLoader
from osgeo import gdal, ogr, osr

from gislite.const import COLOR_INDEX


def get_mts(afile=None):
    '''
    输出最近修改时间
    '''
    if afile:
        return os.path.getmtime(afile)
    else:
        if os.path.exists('mts.log'):
            return os.path.getatime('mts.log')
        else:
            return 0


def lyr_list(xls_file):
    '''
    解析得到excel表内的值放入列表内
    '''
    sheet = load_workbook(filename=xls_file).active

    max_row_num = sheet.max_row

    out_str = []
    for row in range(1, max_row_num + 1):
        the_cell = 'maplet_' + sheet.cell(row=row, column=1).value
        out_str.append(the_cell)
    return out_str


def get_html_title(html_file):
    '''
    Get the title of the page.
    '''
    return BeautifulSoup(open(html_file), "html.parser").title.text


def render_html(tmpl, outfile, **kwargs):
    '''
    Render HTML file using Jinja2.
    '''
    jinja_env = Environment(loader=FileSystemLoader('templates'))
    template = jinja_env.get_template(tmpl)
    with open(outfile, "w") as fh:
        fh.write(template.render(kwargs))


def hex2dec(string_num):
    '''
    Convert HEX to DEC.
    '''
    return str(int(string_num.upper(), 16))


def xlsx2dict(xls_file):
    '''
    将 XLSX 文件中记录的信息转换为 Python dict.
    '''

    sheet = load_workbook(filename=xls_file).active

    max_row_num = sheet.max_row
    max_col_num = sheet.max_column
    out_str = ''
    for row in range(1, max_row_num + 1):

        the_str = ''
        sig = True
        for col in range(1, max_col_num + 1):

            the_cell = sheet.cell(row=row, column=col)
            if the_cell and the_cell.value:

                the_cell_value = the_cell.value

                if isinstance(the_cell_value, str) and the_cell_value.startswith('#'):
                    # 直接定义颜色的情况
                    the_cell_value = '"{}"'.format(the_cell_value.strip())
                else:
                    # 进行颜色判断
                    colors = the_cell.fill.fgColor.index
                    # print(colors)

                    # '00000000' for not filled, 0 for `white`.
                    if colors in ['00000000', 0]:
                        # 无颜色定义
                        pass
                    elif isinstance(colors, int):
                        # 从颜色索引中获取
                        the_cell_value = '"#{}{}"'.format(
                            COLOR_INDEX[colors][2:],
                            COLOR_INDEX[colors][:2]
                        )
                    elif isinstance(colors, str) and len(colors) == 8:
                        # red = int(hex2dec(colors[2:4]))
                        # green = int(hex2dec(colors[4:6]))
                        # blue = int(hex2dec(colors[6:8]))
                        # the_cell_value = [red, green, blue]
                        the_cell_value = '"#{}{}"'.format(colors[2:], colors[:2])

                mf_keys = ['class',
                           'classitem',
                           'labelitem',
                           'data',
                           'labelminscaledenom',
                           'labelmaxscaledenom',
                           'encoding',
                           'processing', ]
                if str(the_cell_value).lower() in mf_keys:
                    the_str = '- ' + the_str

                if sig:
                    the_str = the_str + str(the_cell_value) + ': '
                    sig = False
                else:
                    the_str = the_str + str(the_cell_value)
            else:
                the_str = the_str + '  '
            # print(the_str)
        out_str = out_str + the_str + '\r'

    with open('xx_out.xbj', 'w') as fo:
        fo.write(out_str)

    return yaml.load(out_str)


def get_epsg_code(img_file, raster=False):
    '''
    获取 EPSG 代码。
    '''
    # print(img_file)
    if os.path.isdir(img_file):
        raster = True
    elif img_file.lower().endswith('.tif'):
        raster = True
    if raster:
        # print(img_file)
        gdal_open = gdal.Open(img_file)
        srs = gdal_open.GetProjection()

        sr2 = osr.SpatialReference()
        sr2.SetFromUserInput(srs)
        # print(sr2.ExportToPrettyWkt())

        return {'epsg_code': '',
                'proj4_code': sr2.ExportToProj4(),
                'geom_type': 'raster'}
    else:
        ds = ogr.Open(img_file)
        lyr = ds.GetLayer(0)
        srs = lyr.GetSpatialRef()

        # sr2 = osr.SpatialReference()
        # sr2.SetFromUserInput(srs)
        # epsg_code = srs.GetAttrValue("AUTHORITY", 1)
        # else:
        #     epsg_code = '4326'
        # if epsg_code:
        #     pass
        # else:
        #     epsg_code = '4326'
        geom = None
        idx = 0
        while not geom:
            feat = lyr.GetFeature(idx)
            geom = feat.GetGeometryRef()
            idx = idx + 1
        geom_type = geom.GetGeometryName()

        return {
            'proj4_code': srs.ExportToProj4(),
            # 'epsg_code': '', #epsg_code,
            'geom_type': geom_type}


def rst_for_chapter(secws):
    '''
    '''
    sec_list = os.listdir(secws)
    sec_list = [x for x in sec_list if
                (x.startswith('sec')
                 and not x.endswith('_files')
                 and (x[-3:] not in ['jpg', 'gif', 'png']))]
    sec_list.sort()

    rst_new_list = []
    for sec_dir in sec_list:
        rst_new_list.append(sec_dir)

    idxfile = os.path.join(secws, 'chapter.rst')
    if os.path.exists(idxfile):
        pass
    else:
        with open(idxfile, 'w') as fo:
            fo.write('''Chapter
==============================================

''')
    sec_cnt = open(idxfile).readlines()

    with open(idxfile, 'w') as fo:
        for uu in sec_cnt:
            if '.. toctree::' in uu:
                break
            else:
                fo.write(uu)
        fo.write('''.. toctree::\n   :maxdepth: 2\n\n''')
        for x in rst_new_list:
            fo.write('   {0}\n'.format(x))


def rst_for_book(secws):
    sec_list = os.listdir(secws)
    sec_list = [x for x in sec_list if x[:2] in ['ch', 'pt']]
    sec_list.sort()

    # print(sec_list)

    rst_new_list = []
    for sec_dir in sec_list:
        rst_new_list.append(sec_dir)

    if os.path.exists(os.path.join(secws, 'index.rst')):
        pass
    else:
        return False
    sec_cnt = open(os.path.join(secws, 'index.rst')).readlines()

    # print(rst_new_list)
    with open(os.path.join(secws, 'index.rst'), 'w') as fo:
        for uu in sec_cnt:
            if '.. toctree::' in uu:
                break
            else:
                fo.write(uu)

        if rst_new_list[0].startswith('ch'):
            fo.write('''.. toctree::\n   :maxdepth: 3\n   :numbered: 3\n\n''')
        else:
            fo.write('''.. toctree::\n\n''')

        for x in rst_new_list:
            if x.startswith('ch'):
                fo.write('   {0}/chapter\n'.format(x))
            else:
                fo.write('   {0}/part\n'.format(x))


def clean_sphinx(fuws):
    '''
    do, one by one.
    '''

    for wroot, wdirs, wfiles in os.walk(fuws):
        for wdir in wdirs:
            if wdir.startswith('ch'):
                inws = os.path.join(wroot, wdir)
                rst_for_chapter(inws)

    rst_for_book(fuws)
